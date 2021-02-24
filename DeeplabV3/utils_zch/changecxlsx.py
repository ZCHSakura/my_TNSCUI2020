# coding:utf-8
import openpyxl
import numpy as np


# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
wb.create_sheet('valid')
# 获取对应sheet
ws = wb["valid"]
ws.cell(1, 1).value = 'ID'
ws.cell(1, 2).value = 'CATE'
for i in range(3000, 3644):
    for j in range(2):
        # data = ('{0}_{1}.bmp'.format(i, j + 1), )
        data = ('{0}_{1}'.format(i, j + 1), np.random.randint(0, 2))
        ws.append(data)  # 每次写入一行

wb.save('G:/graduation project/TN-SCUI2020/my_code/augtrain/all_id_cate.xlsx')
